import sys
import types
from pathlib import Path
import os

PHOTSIM_ROOT = os.environ.get(
    "PHOTSIM7_ROOT",
    os.environ.get("PHOTSIM6FT_ROOT", "/home/cxgao/ET/Photosim7"),
)

# 让 Python 能找到同事的源码目录（但不需要安装到环境）
if PHOTSIM_ROOT not in sys.path:
    sys.path.insert(0, PHOTSIM_ROOT)

# 避免当前会话里已经加载过旧的 photsim6（很关键）
for m in list(sys.modules):
    if m == "photsim6" or m.startswith("photsim6."):
        sys.modules.pop(m)

# 重要：不要 `import photsim7` 顶层包。
# 因为 photsim7/__init__.py 会导入 dashboard/dash；如果你的环境没装 dash，会直接崩。
# 我们在这里创建一个“轻量包别名” photsim6，把它的搜索路径指向当前 Photosim 源码目录：
# 这样 `import photsim6.simulator` 等会正常加载，但不会执行 photsim7/__init__.py。
_photsim_root_path = Path(PHOTSIM_ROOT)
_candidate_src_dirs = [
    _photsim_root_path / "photsim7",
    _photsim_root_path / "photsim6ft",
]
PHOTSIM6_SRC_DIR = str(
    next((path for path in _candidate_src_dirs if path.exists()), _candidate_src_dirs[0])
)
photsim6_alias_pkg = types.ModuleType("photsim6")
photsim6_alias_pkg.__path__ = [PHOTSIM6_SRC_DIR]
photsim6_alias_pkg.__package__ = "photsim6"
sys.modules["photsim6"] = photsim6_alias_pkg

# photsim simulator.py 会 import dashboard.py；而 dashboard.py 依赖 dash。
# 本脚本始终使用 start_ray=False，不需要 dashboard 真正工作。
# 若当前环境未安装 dash，则注入一个最小 stub，避免 import 时崩溃。
try:
    import dash  # type: ignore  # noqa: F401
except Exception:
    dash_stub = types.ModuleType("dash")

    def _dash_component(*args, **kwargs):
        return {"args": args, "kwargs": kwargs}

    class _DashNamespace:
        def __getattr__(self, name):
            return _dash_component

    class _DummyDashApp:
        def __init__(self, *args, **kwargs):
            self.server = None
            self.layout = None
            self.index_string = ""

        def callback(self, *args, **kwargs):
            def _decorator(func):
                return func

            return _decorator

        def run(self, *args, **kwargs):
            return None

    dash_stub.Dash = _DummyDashApp
    dash_stub.html = _DashNamespace()
    dash_stub.dcc = _DashNamespace()
    dash_stub.dash_table = _DashNamespace()
    sys.modules["dash"] = dash_stub

    dash_dependencies = types.ModuleType("dash.dependencies")
    dash_dependencies.Input = _dash_component
    dash_dependencies.Output = _dash_component
    sys.modules["dash.dependencies"] = dash_dependencies

    dash_exceptions = types.ModuleType("dash.exceptions")

    class _PreventUpdate(Exception):
        pass

    dash_exceptions.PreventUpdate = _PreventUpdate
    sys.modules["dash.exceptions"] = dash_exceptions

import subprocess
from openpyxl import load_workbook

os.environ["ET_DATA_DIR"] = os.environ.get("ET_DATA_DIR", r"/home/cxgao/ET/Photosim6/data")

import matplotlib.pyplot as plt
import numpy as np
from astropy import units as u
import json
import random

import torch


# -----------------------------------------------------------------------------
# Minimal replacements for symbols previously provided by photsim6 common imports

opj = os.path.join

try:
    from tqdm.auto import trange  # Notebook/terminal friendly
except Exception:  # pragma: no cover
    trange = range


# --------------------------------------------------------------
# Import commonly used packages
# from photsim6.config import common_import_script
# %run $common_import_script

# Import managers
from photsim6.simulator import Simulator
from photsim6.configurator import ConfigurationManager
from photsim6.time import resolve_detector_frame_timing
from photsim6.variants import VariantManager

# Import additional utilities
from photsim6.plot import (
    plot_jitter_integrated_psf_models,
    plot_single_frame_tpf,
    plot_gif_tpfs,
    plot_base_psf_models,
    det_motion_and_component_plot,
    meas_jit_drift_rms,
    plot_psd,
)
from photsim6.data_generators import (
    real_freqs,
    real_mags,
    psd_to_motion,
    gen_psd_motion,
    fft_bins_to_name_str,
    fft_bins_to_fullname_str,
    gen_power_spec,
    load_power_spectrum_config,
    create_power_spec_components,
    generate_jitter,
)
from photsim6.config import BASE_DATA_DIR

import pickle


ET_SINGLE_SCOPE_ZEROPOINT_E_PER_S = 615.75 * 1_961_225


def surface_brightness_mag_per_arcsec2_to_flux_per_pix(
    mag_per_arcsec2: float, pixel_scale_arcsec_per_pix: float
) -> float:
    """Convert ET surface brightness to per-pixel electron rate.

    The conversion uses the same ET zero point already hard-coded in
    `photsim6ft.field.Stars.build_catalog()` to keep stars and diffuse
    background on a consistent flux scale.
    """

    pixel_area_arcsec2 = pixel_scale_arcsec_per_pix**2
    return float(
        ET_SINGLE_SCOPE_ZEROPOINT_E_PER_S
        * 10 ** (-0.4 * mag_per_arcsec2)
        * pixel_area_arcsec2
    )


# ============================================================================
# 导星探测器参数（由用户提供）
# ============================================================================
GUIDE_PIXEL_SCALE_ARCSEC_PER_PIX = 3.146
GUIDE_PIXEL_WIDTH_UM = 6.5
GUIDE_FOV_DEG = 1.7
GUIDE_EXPOSURE_S = 0.25
GUIDE_READOUT_S = 0.0
GUIDE_OBSERVING_DURATION_S = 6.0
GUIDE_READOUT_NOISE_ADU_PER_PIX = 5.0
GUIDE_DARK_PLUS_SCATTERED_E_PER_S_PER_PIX = 10.0
# 用户只给了暗电流+杂散光总和；两者都开启时，任意拆分都等价于同一个总泊松过程。
GUIDE_DARK_CURRENT_E_PER_S_PER_PIX = GUIDE_DARK_PLUS_SCATTERED_E_PER_S_PER_PIX / 2.0
GUIDE_SCATTERED_LIGHT_E_PER_S_PER_PIX = GUIDE_DARK_PLUS_SCATTERED_E_PER_S_PER_PIX / 2.0
GUIDE_INTER_PRV_RMS_PERCENT = 3.0
GUIDE_BACKGROUND_SURFACE_BRIGHTNESS_MAG_PER_ARCSEC2 = 22.0
GUIDE_BACKGROUND_FLUX_E_PER_S_PER_PIX = (
    surface_brightness_mag_per_arcsec2_to_flux_per_pix(
        GUIDE_BACKGROUND_SURFACE_BRIGHTNESS_MAG_PER_ARCSEC2,
        GUIDE_PIXEL_SCALE_ARCSEC_PER_PIX,
    )
)
GUIDE_PSF_FIELD_ANGLE_DEG = 14.0
GUIDE_PSF_FIELD_ID = 7


# ============================================================================
# 用户可改的“导星探测器仿真”关键设置（本脚本层实现，不改 photsim6ft 包）
# ============================================================================
# 1) 单个导星探测器视场。
FOV_DEG = GUIDE_FOV_DEG

# 1.5) 4 个导星探测器的天区中心接口。
FIELD_CENTERS_DEG = [
    (278.18437, 37.57037),
    (269.58206, 57.89966),
    (310.62391, 59.26764),
    (305.03563, 38.47553),
]

# 是否一次性顺序跑完多个 batch。
# - True：当前脚本会自动按 BATCH_INDICES_TO_RUN 逐个拉起子进程。
# - False：仅运行当前 FIELD_CENTER_INDEX 指定的单个 batch。
RUN_ALL_BATCHES = True

# 指定需要运行的 batch 索引；None 表示跑完 FIELD_CENTERS_DEG 中的全部中心。
# 例如只跑 batch0 和 batch2，可改成 [0, 2]。
BATCH_INDICES_TO_RUN = None

# 当前批次索引：0/1/2/3
FIELD_CENTER_INDEX = 0

_FIELD_CENTER_INDEX_ENV = os.environ.get("ET_FIELD_CENTER_INDEX")
if _FIELD_CENTER_INDEX_ENV is not None:
    FIELD_CENTER_INDEX = int(_FIELD_CENTER_INDEX_ENV)

IS_CHILD_BATCH_RUN = os.environ.get("ET_CHILD_BATCH_RUN") == "1"
_RUN_ALL_BATCHES_ENV = os.environ.get("ET_RUN_ALL_BATCHES")
if _RUN_ALL_BATCHES_ENV is not None:
    RUN_ALL_BATCHES = _RUN_ALL_BATCHES_ENV.strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
        "on",
    }


def run_all_batches_sequentially() -> None:
    batch_indices = (
        list(range(len(FIELD_CENTERS_DEG)))
        if BATCH_INDICES_TO_RUN is None
        else [int(idx) for idx in BATCH_INDICES_TO_RUN]
    )

    invalid = [idx for idx in batch_indices if not (0 <= idx < len(FIELD_CENTERS_DEG))]
    if invalid:
        raise ValueError(
            f"BATCH_INDICES_TO_RUN 中存在越界索引 {invalid}；"
            f"合法范围应为 [0, {len(FIELD_CENTERS_DEG) - 1}]。"
        )

    script_path = str(Path(__file__).resolve())
    for batch_index in batch_indices:
        batch_ra_deg, batch_dec_deg = FIELD_CENTERS_DEG[batch_index]
        print(
            f"[Batch launcher] start batch={batch_index}  "
            f"RA={batch_ra_deg:.6f} deg  DEC={batch_dec_deg:.6f} deg"
        )
        env = os.environ.copy()
        env["ET_CHILD_BATCH_RUN"] = "1"
        env["ET_FIELD_CENTER_INDEX"] = str(batch_index)
        subprocess.run([sys.executable, script_path], env=env, check=True)


if RUN_ALL_BATCHES and not IS_CHILD_BATCH_RUN:
    run_all_batches_sequentially()
    sys.exit(0)

if not (0 <= FIELD_CENTER_INDEX < len(FIELD_CENTERS_DEG)):
    raise ValueError(
        f"FIELD_CENTER_INDEX={FIELD_CENTER_INDEX} 越界；"
        f"请确保它在 [0, {len(FIELD_CENTERS_DEG) - 1}] 内。"
    )

FIELD_CENTER_RA_DEG, FIELD_CENTER_DEC_DEG = FIELD_CENTERS_DEG[FIELD_CENTER_INDEX]
print(
    f"[Sky center] batch={FIELD_CENTER_INDEX}  RA={FIELD_CENTER_RA_DEG:.6f} deg  DEC={FIELD_CENTER_DEC_DEG:.6f} deg"
)

# 2) drift-jitter 分界：以“单个 raw frame 的有效积分时间”作为时间尺度。
#    - CCD：积分窗口 = Exposure Duration * Simulation Cadence Mult
#    - CMOS：积分窗口 = (Exposure + Readout) * Simulation Cadence Mult
#      因为当前模型假设 CMOS 在读出期间仍持续积光。
#    解释：把短于该积分窗口的运动平均进 PSF（jitter）；更慢的变化保留为帧间质心漂移（drift）。
#    因此主探测器 10s 与导星探测器 250ms 会自动得到不同 split_hz。
USE_CADENCE_SPLIT = True

# 2.1) 可选：手动覆盖 split_hz（一般不需要）
# - None：自动 split_hz = 1 / raw_frame_integration_s（推荐）
# - 数值：强制使用该 split_hz
SPLIT_HZ_OVERRIDE = None

# 3) 导星星表查询深度：
# - 先从 Gaia G=11 开始查；
# - 若该视场内星数不足目标值，则逐步放宽到更暗的星；
# - 只把最亮的若干颗送进仿真，避免逐星 PSF 叠加过慢。
GAIA_GMAG_LIM = 11.0
GAIA_GMAG_LIM_STEP = 0.5
GAIA_GMAG_LIM_MAX = 16.0
GUIDE_STAR_QUERY_BACKEND = "et_focalplane"
GUIDE_GAIA_CATALOG_DIR = os.environ.get("GUIDE_GAIA_CATALOG_DIR", "/home/cxgao/gaia_dr3_19mag")
ET_FOCALPLANE_ROOT = os.environ.get("ET_FOCALPLANE_ROOT", "/home/cxgao/ET/et_focalplane")
ET_FOCALPLANE_DATA_DIR = str(Path(ET_FOCALPLANE_ROOT) / "data")
ET_FOCALPLANE_SRC_DIR = str(Path(ET_FOCALPLANE_ROOT) / "src")
GUIDE_QUERY_TARGET_EPOCH = 2000.0
GUIDE_QUERY_CROP_TO_SIM_FRAME = True
GUIDE_QUERY_CROP_MARGIN_PIX = 0.0

# 3.1) 批次级 field polar angle（单位：度）。
# DVA / thermal drift 在当前裁剪 ET workflow 中按“探测器级共享方向”近似处理，
# 这里显式固定为导星探测器对应的 14 deg PSF 方向，并在图像渲染与 truth 导出中复用。
FIELD_POLAR_ANGLE_DEG = GUIDE_PSF_FIELD_ANGLE_DEG
FIELD_POLAR_ANGLE_RAD = float(np.deg2rad(FIELD_POLAR_ANGLE_DEG))

# 3.2) 星数上限（重要的性能开关）
# photsim6ft 的 `generate_background_starlight_frames` 会对每颗星逐个生成 PSF 并累加，
# 星数一旦上千，每帧耗时会非常夸张，GPU 上还可能触发 watchdog timeout。
# 为避免大视场下逐星 PSF 叠加过慢，这里默认只保留最亮的若干颗星（包含 target 星）。
MAX_SIM_STARS = 200

# 3.2.1) 导星星表选择规则：
# - 每个 batch 目标至少 150 颗星
# - 若 11 等内不足，则自动放宽最暗星限制，直到凑够目标值或达到查询上限
# - 最终仍受 MAX_SIM_STARS 限制，默认保留最亮的 200 颗（包含 target 星）
GUIDE_MIN_CATALOG_STARS = 150

# 3.3) 是否对整幅星场施加静态亚像元偏移。
# - False：默认关闭，便于把星表真值与像面真值严格对齐，适合导星精度评估。
# - True：为整幅星场加入同一个 (dx, dy) 常量偏移，模拟天空相对像元网格的静态相位。
#   若不手动指定 offset，脚本会在 [-0.5, 0.5] pix 内随机采样，并写入 run_meta.json。
APPLY_STATIC_FIELD_OFFSET = False
STATIC_FIELD_OFFSET_X_PIX = None
STATIC_FIELD_OFFSET_Y_PIX = None

# 3.5) 计算设备：你希望使用 GPU。
# - 设为 "cuda" 会尝试使用 GPU；若当前 PyTorch/驱动不可用，会自动回退到 CPU。
# - 也可手动设为 "cpu" 强制走 CPU。
REQUESTED_COMPUTE_DEVICE = "cuda"

# 4) 产物输出根目录（你可指定到任意磁盘/目录）：
#    你希望“所有产品”都落在一个确定位置。这里的 OUTPUT_ROOT 就是那个最上层目录。
#    脚本会在其下创建更细的子目录结构（你说子目录怎么组织无所谓）。
OUTPUT_ROOT = "/home/cxgao/ET/FSG_guide_sims"

# 本次运行的名字（目录名的一部分），建议你按项目/日期/参数命名，便于检索。
OUTPUT_RUN_NAME = "guide_det_v1_noise_psf_6s"

# 是否启用流式落盘：每帧（或极小批次）生成后立刻写入 .npz，避免内存累计。
STREAM_SAVE_FRAMES = True


# 每次生成并保存多少帧。
# - 1：最安全（逐帧保存），几乎不可能因“堆帧”而爆内存/显存。
# - >1：更快，但会短暂占用更多内存（尤其是 subpixel 阶段）。
FRAME_BATCH_SIZE = 1

# 5) 随机序列一致性：跨不同 FIELD_CENTER_INDEX 仍保持同一套随机序列
#    目标：保证 drift/jitter、MD 随机游走、以及所有依赖 RNG 的 time series 在不同天区中心下完全一致。
GLOBAL_SEED = 12345

# 建议保持 True：每个 batch 在生成 dynamic params 前重置种子，避免“星场查询/随机偏移”消耗 RNG 影响 motion。
RESEED_BEFORE_DYNAMIC_PARAMS = True


def set_global_seed(seed: int) -> None:
    """设置所有常见 RNG 的随机种子，保证跨批次可复现。"""
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def env_flag(name: str, default: bool) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    value = value.strip().lower()
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    raise ValueError(f"环境变量 {name}={value!r} 不是合法布尔值。")


def env_int(name: str, default=None):
    value = os.environ.get(name)
    if value is None or value == "":
        return default
    return int(value)


def env_str(name: str, default: str) -> str:
    value = os.environ.get(name)
    if value is None:
        return default
    value = value.strip()
    return value if value else default


def resolve_field_offsets(
    apply_offset: bool,
    offset_x_pix,
    offset_y_pix,
) -> tuple[float, float]:
    """Resolve a single static field offset and reuse it across adaptive Gaia retries."""

    if apply_offset:
        resolved_x = (
            float(offset_x_pix)
            if offset_x_pix is not None
            else float(np.random.uniform(-0.5, 0.5))
        )
        resolved_y = (
            float(offset_y_pix)
            if offset_y_pix is not None
            else float(np.random.uniform(-0.5, 0.5))
        )
        return resolved_x, resolved_y

    return (
        0.0 if offset_x_pix is None else float(offset_x_pix),
        0.0 if offset_y_pix is None else float(offset_y_pix),
    )


def _slice_star_data_by_indices(star_data: dict, keep_indices) -> dict:
    """Return a copy of star_data keeping only per-star entries at keep_indices."""

    keep = np.asarray(keep_indices, dtype=int)
    n_stars = int(len(np.asarray(star_data["x0"])))
    sliced = {}

    for key, value in star_data.items():
        if isinstance(value, str):
            sliced[key] = value
            continue

        if np.isscalar(value):
            sliced[key] = value
            continue

        try:
            arr = np.asarray(value)
        except Exception:
            sliced[key] = value
            continue

        if arr.ndim >= 1 and arr.shape[0] == n_stars:
            sliced[key] = arr[keep].copy()
        else:
            sliced[key] = value

    return sliced


def enrich_star_data_truth_columns(star_data: dict) -> dict:
    """Add detector-absolute truth columns when the et_focalplane backend provides them."""

    enriched = dict(star_data)
    field_offset_x = float(enriched.get("field_offset_x_pix", 0.0))
    field_offset_y = float(enriched.get("field_offset_y_pix", 0.0))

    if "detector_xpix" in enriched:
        enriched["detector_xpix_shifted"] = (
            np.asarray(enriched["detector_xpix"], dtype=float) + field_offset_x
        )
    if "detector_ypix" in enriched:
        enriched["detector_ypix_shifted"] = (
            np.asarray(enriched["detector_ypix"], dtype=float) + field_offset_y
        )
    if "target_center_xpix" in enriched:
        enriched["target_center_xpix_shifted"] = (
            float(enriched["target_center_xpix"]) + field_offset_x
        )
    if "target_center_ypix" in enriched:
        enriched["target_center_ypix_shifted"] = (
            float(enriched["target_center_ypix"]) + field_offset_y
        )

    return enriched


def select_guide_star_indices(
    star_df,
    min_catalog_stars: int,
    max_sim_stars: int | None,
):
    """Select the brightest guide stars, preserving target star at index 0."""

    mags = np.asarray(star_df["Kepler Mag"], dtype=float)
    if mags.size == 0:
        raise ValueError("空星表无法用于导星仿真。")

    candidate_idx = np.arange(mags.size, dtype=int)

    # 目标星始终保留在第 0 行，避免破坏 photsim6ft 对 target/background 的约定。
    keep = [0]
    for idx in candidate_idx[np.argsort(mags[candidate_idx])].tolist():
        if idx == 0:
            continue
        keep.append(idx)

    n_available = len(keep)

    if max_sim_stars is not None and max_sim_stars > 0 and len(keep) > max_sim_stars:
        keep = keep[: int(max_sim_stars)]

    return {
        "keep_indices": np.asarray(keep, dtype=int),
        "mags_all": mags,
        "n_available": int(n_available),
        "n_selected": int(len(keep)),
        "selected_faintest_mag": float(np.max(mags[np.asarray(keep, dtype=int)])),
        "meets_min_count": bool(n_available >= int(min_catalog_stars)),
    }


def query_guide_star_field_adaptive(
    *,
    ra,
    dec,
    px_cnt,
    px_scale,
    apply_offset: bool,
    plot: bool,
    offset_x_pix,
    offset_y_pix,
    base_gaia_gmag_lim: float,
    min_catalog_stars: int,
    gaia_gmag_step: float,
    gaia_gmag_lim_max: float,
    query_backend: str,
    catalog_dir: str | None,
    et_focalplane_data_dir: str | None,
    et_focalplane_src_dir: str | None,
    target_epoch: float,
    crop_to_simulation_frame: bool,
    crop_margin_pix: float,
):
    """Query Gaia repeatedly, relaxing the magnitude limit until enough stars are found."""

    resolved_offset_x_pix, resolved_offset_y_pix = resolve_field_offsets(
        apply_offset=apply_offset,
        offset_x_pix=offset_x_pix,
        offset_y_pix=offset_y_pix,
    )

    tried_limits = []
    gaia_gmag_lim = float(base_gaia_gmag_lim)
    star_data = None
    backend = query_backend.strip().lower()

    while True:
        if backend == "et_focalplane":
            star_data = mk_real_field_stars_et_focalplane(
                target_ra=ra,
                target_dec=dec,
                catalog_dir=catalog_dir or GUIDE_GAIA_CATALOG_DIR,
                registry_data_dir=et_focalplane_data_dir or ET_FOCALPLANE_DATA_DIR,
                px_cnt=px_cnt,
                apply_offset=apply_offset,
                mag_lim=gaia_gmag_lim,
                plot=False,
                offset_x_pix=resolved_offset_x_pix,
                offset_y_pix=resolved_offset_y_pix,
                target_epoch=target_epoch,
                crop_to_simulation_frame=crop_to_simulation_frame,
                crop_margin_pix=crop_margin_pix,
                et_focalplane_src=et_focalplane_src_dir or ET_FOCALPLANE_SRC_DIR,
            )
        elif backend == "legacy_box_gaia":
            star_data = mk_real_field_stars_2(
                ra=ra,
                dec=dec,
                px_cnt=px_cnt,
                px_scale=px_scale,
                apply_offset=apply_offset,
                mag_lim=gaia_gmag_lim,
                plot=False,
                offset_x_pix=resolved_offset_x_pix,
                offset_y_pix=resolved_offset_y_pix,
            )
        else:
            raise ValueError(
                f"Unsupported GUIDE_STAR_QUERY_BACKEND={query_backend!r}. "
                "Use 'et_focalplane' or 'legacy_box_gaia'."
            )

        n_found = int(len(np.asarray(star_data["x0"])))
        tried_limits.append(
            {
                "gaia_gmag_lim": float(gaia_gmag_lim),
                "n_stars_found": n_found,
            }
        )

        if n_found >= int(min_catalog_stars) or gaia_gmag_lim >= float(
            gaia_gmag_lim_max
        ):
            break

        gaia_gmag_lim = min(
            float(gaia_gmag_lim_max),
            float(gaia_gmag_lim + gaia_gmag_step),
        )

    if plot:
        if backend == "et_focalplane":
            star_data = mk_real_field_stars_et_focalplane(
                target_ra=ra,
                target_dec=dec,
                catalog_dir=catalog_dir or GUIDE_GAIA_CATALOG_DIR,
                registry_data_dir=et_focalplane_data_dir or ET_FOCALPLANE_DATA_DIR,
                px_cnt=px_cnt,
                apply_offset=apply_offset,
                mag_lim=gaia_gmag_lim,
                plot=True,
                offset_x_pix=resolved_offset_x_pix,
                offset_y_pix=resolved_offset_y_pix,
                target_epoch=target_epoch,
                crop_to_simulation_frame=crop_to_simulation_frame,
                crop_margin_pix=crop_margin_pix,
                et_focalplane_src=et_focalplane_src_dir or ET_FOCALPLANE_SRC_DIR,
            )
        else:
            star_data = mk_real_field_stars_2(
                ra=ra,
                dec=dec,
                px_cnt=px_cnt,
                px_scale=px_scale,
                apply_offset=apply_offset,
                mag_lim=gaia_gmag_lim,
                plot=True,
                offset_x_pix=resolved_offset_x_pix,
                offset_y_pix=resolved_offset_y_pix,
            )

    return star_data, {
        "query_backend": backend,
        "query_detector_id": str(star_data.get("detector_id", "")),
        "query_target_epoch": float(target_epoch),
        "query_base_gaia_gmag_lim": float(base_gaia_gmag_lim),
        "query_final_gaia_gmag_lim": float(gaia_gmag_lim),
        "query_gaia_gmag_step": float(gaia_gmag_step),
        "query_gaia_gmag_lim_max": float(gaia_gmag_lim_max),
        "query_crop_to_simulation_frame": bool(crop_to_simulation_frame),
        "query_crop_margin_pix": float(crop_margin_pix),
        "query_history": tried_limits,
        "query_reached_min_star_target": bool(
            len(np.asarray(star_data["x0"])) >= int(min_catalog_stars)
        ),
        "resolved_field_offset_x_pix": float(resolved_offset_x_pix),
        "resolved_field_offset_y_pix": float(resolved_offset_y_pix),
    }


def build_variant_settings(**overrides):
    settings = dict(
        enable_stellar_photon_noise=False,
        enable_background_light=True,
        enable_scattered_light=True,
        enable_dark_current=True,
        enable_readout_noise=True,
        enable_gain=False,
        enable_target_star=True,
        enable_background_stars=True,
        enable_jitter=False,
        enable_dva_drift=False,
        enable_pointing_drift=False,
        enable_psf_breathing=False,
        enable_inter_pixel_response=False,
        enable_intra_pixel_response=False,
        enable_pixel_phase_response=False,
    )
    settings.update(overrides)
    return settings


EFFECT_PROFILES = {
    "full": dict(
        name="full",
        description="full",
        variant_settings={},
        include_pointing_drift=True,
        include_dva=True,
        include_thermal=True,
        include_momentum_dump=True,
        uses_jitter_integrated_psf=True,
    ),
    "v1_noise_psf": dict(
        name="v1_noise_psf",
        description="v1_noise_psf",
        variant_settings=build_variant_settings(),
        include_pointing_drift=False,
        include_dva=False,
        include_thermal=False,
        include_momentum_dump=False,
        uses_jitter_integrated_psf=False,
    ),
    "v2_point_drift_jitter": dict(
        name="v2_point_drift_jitter",
        description="v2_point_drift_jitter",
        variant_settings=build_variant_settings(
            enable_jitter=True,
            enable_pointing_drift=True,
        ),
        include_pointing_drift=True,
        include_dva=False,
        include_thermal=False,
        include_momentum_dump=False,
        uses_jitter_integrated_psf=True,
    ),
    "v3_dva": dict(
        name="v3_dva",
        description="v3_dva",
        variant_settings=build_variant_settings(
            enable_dva_drift=True,
        ),
        include_pointing_drift=False,
        include_dva=True,
        include_thermal=False,
        include_momentum_dump=False,
        uses_jitter_integrated_psf=False,
    ),
    "v4_thermal": dict(
        name="v4_thermal",
        description="v4_thermal",
        variant_settings=build_variant_settings(
            enable_pointing_drift=True,
        ),
        include_pointing_drift=False,
        include_dva=False,
        include_thermal=True,
        include_momentum_dump=False,
        uses_jitter_integrated_psf=False,
    ),
    "v5_prv_subpixel": dict(
        name="v5_prv_subpixel",
        description="v5_prv_subpixel",
        variant_settings=build_variant_settings(
            enable_inter_pixel_response=True,
            enable_intra_pixel_response=True,
            enable_pixel_phase_response=True,
        ),
        include_pointing_drift=False,
        include_dva=False,
        include_thermal=False,
        include_momentum_dump=False,
        uses_jitter_integrated_psf=False,
    ),
}


EFFECT_PROFILE_ALIASES = {
    "full": "full",
    "version1": "v1_noise_psf",
    "v1": "v1_noise_psf",
    "v1_noise_psf": "v1_noise_psf",
    "version2": "v2_point_drift_jitter",
    "v2": "v2_point_drift_jitter",
    "v2_point_drift_jitter": "v2_point_drift_jitter",
    "version3": "v3_dva",
    "v3": "v3_dva",
    "v3_dva": "v3_dva",
    "version4": "v4_thermal",
    "v4": "v4_thermal",
    "v4_thermal": "v4_thermal",
    "version5": "v5_prv_subpixel",
    "v5": "v5_prv_subpixel",
    "v5_prv_subpixel": "v5_prv_subpixel",
}


requested_effect_profile = env_str("ET_EFFECT_PROFILE", "v1_noise_psf").lower()
effect_profile_key = EFFECT_PROFILE_ALIASES.get(
    requested_effect_profile, requested_effect_profile
)
if effect_profile_key not in EFFECT_PROFILES:
    valid_profiles = ", ".join(sorted(EFFECT_PROFILES))
    raise ValueError(
        f"未知 ET_EFFECT_PROFILE={requested_effect_profile!r}。"
        f" 可选值: {valid_profiles}"
    )

ACTIVE_EFFECT_PROFILE = EFFECT_PROFILES[effect_profile_key]
PROFILE_TARGET_FRAMES = env_int("ET_PROFILE_TARGET_FRAMES", None)

RUN_ALL_BATCHES = env_flag("ET_RUN_ALL_BATCHES", RUN_ALL_BATCHES)
OUTPUT_ROOT = env_str("ET_OUTPUT_ROOT_OVERRIDE", OUTPUT_ROOT)
OUTPUT_RUN_NAME = env_str("ET_OUTPUT_RUN_NAME_OVERRIDE", OUTPUT_RUN_NAME)
MAX_SIM_STARS = env_int("ET_MAX_SIM_STARS", MAX_SIM_STARS)

print(
    "[Effect profile] "
    f"{ACTIVE_EFFECT_PROFILE['name']}  "
    f"(pointing={ACTIVE_EFFECT_PROFILE['include_pointing_drift']}, "
    f"jitter={ACTIVE_EFFECT_PROFILE['variant_settings'].get('enable_jitter', True)}, "
    f"dva={ACTIVE_EFFECT_PROFILE['include_dva']}, "
    f"thermal={ACTIVE_EFFECT_PROFILE['include_thermal']}, "
    f"prv={ACTIVE_EFFECT_PROFILE['variant_settings'].get('enable_inter_pixel_response', True) or ACTIVE_EFFECT_PROFILE['variant_settings'].get('enable_intra_pixel_response', True) or ACTIVE_EFFECT_PROFILE['variant_settings'].get('enable_pixel_phase_response', True)})"
)


def get_output_dirs():
    """构造并创建本次运行的输出目录。

    返回：
    - run_dir：本次运行根目录
    - batch_dir：当前 FIELD_CENTER_INDEX 的目录
    - frames_dir：逐帧（或逐小批次）保存 .npz 的目录
    - cache_dir：缓存目录（例如 jitter 缓存）
    """
    run_dir = opj(OUTPUT_ROOT, OUTPUT_RUN_NAME)
    batch_dir = opj(
        run_dir,
        f"batch{FIELD_CENTER_INDEX}_ra{FIELD_CENTER_RA_DEG:.4f}_dec{FIELD_CENTER_DEC_DEG:.4f}",
    )
    frames_dir = opj(batch_dir, "frames")
    cache_dir = opj(run_dir, "cache")
    os.makedirs(frames_dir, exist_ok=True)
    os.makedirs(cache_dir, exist_ok=True)
    return run_dir, batch_dir, frames_dir, cache_dir


def pick_compute_device(requested: str) -> str:
    requested = (requested or "").strip().lower()
    if requested == "cpu":
        return "cpu"
    if requested == "cuda":
        return "cuda" if torch.cuda.is_available() else "cpu"
    # 兜底：未知字符串时，按“有 GPU 用 GPU”处理
    return "cuda" if torch.cuda.is_available() else "cpu"


def build_runtime_config_xlsx(source_path: str, batch_dir: str) -> str:
    """Create a per-run config copy with hidden telescope FOV offset disabled.

    We intentionally keep the source Excel untouched. Older config sheets may still
    carry `Telescope FOV Max Offset > 0`, but the simulator no longer permits that
    hidden random offset. For runtime safety and reproducibility, sanitize it here.
    """

    source = Path(source_path).expanduser().resolve()
    runtime_path = (
        Path(batch_dir) / f"{source.stem}.runtime_no_fov_offset{source.suffix}"
    )

    wb = load_workbook(source)
    replacements = {
        "Telescope FOV Max Offset": (
            0,
            "Must remain 0. Hidden telescope-wide random FOV offset is disabled for "
            "ET guide simulations; use explicit field offsets only when intentionally configured.",
        ),
        "Target Max Offset": (
            0,
            "Must remain 0 for ET guide simulations unless a target-star-only offset is "
            "explicitly required by the experiment design.",
        ),
    }
    patched = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            if len(row) < 5:
                continue
            key = row[1].value
            if key not in replacements:
                continue
            replacement_value, replacement_desc = replacements[key]
            current = row[2].value
            if current not in (
                replacement_value,
                float(replacement_value),
                str(replacement_value),
                None,
            ):
                print(
                    "[Config sanitize] "
                    f"{source.name}: {key} {current} -> {replacement_value}"
                )
            row[2].value = replacement_value
            row[4].value = replacement_desc
            patched.add(str(key))

    missing = [key for key in replacements if key not in patched]
    if missing:
        raise ValueError(
            "未在输入配置表中找到以下参数，无法构造运行时配置副本：" f" {missing}"
        )

    wb.save(runtime_path)
    return str(runtime_path)


# ---- global init: deterministic RNG + output dirs ----
set_global_seed(GLOBAL_SEED)
RUN_DIR, BATCH_DIR, FRAMES_DIR, CACHE_DIR = get_output_dirs()
COMPUTE_DEVICE = pick_compute_device(REQUESTED_COMPUTE_DEVICE)

# 先落一个最小 metadata，后面读取完 xlsx / 推导参数后会再补充写入。
meta_path = opj(BATCH_DIR, "run_meta.json")
if not os.path.isfile(meta_path):
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "field_center_index": FIELD_CENTER_INDEX,
                "field_center_ra_deg": FIELD_CENTER_RA_DEG,
                "field_center_dec_deg": FIELD_CENTER_DEC_DEG,
                "global_seed": GLOBAL_SEED,
                "requested_compute_device": REQUESTED_COMPUTE_DEVICE,
                "compute_device": COMPUTE_DEVICE,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )


def load_pickle(pklf):
    with open(pklf, "rb") as f:
        dd = pickle.load(f)

    return dd


plt.close("all")
# %matplotlib inline

"""
Working with this configuration:
------------------------------------------------------------------------------------------
|                         Last run date: 2025-06-05 02:05 PM EDT                         |
------------------------------------------------------------------------------------------
| Python | Jupyterlab | Numpy | Scipy  | Astropy | Matplotlib | Tensorflow |   Pytorch   |
------------------------------------------------------------------------------------------
| 3.12.9 |   4.3.6    | 2.2.4 | 1.15.2 |  7.0.1  |   3.10.1   |    N/A     | 2.6.0+cu126 |
------------------------------------------------------------------------------------------
"""

# config_xlsx_full_path = 'et_inputs_D2_PSF(241006)_Ji(0.24as)_Dr(0.03as_TESS_to_ET)_MD(limit_0.15as_Cyc_3d)_B(TESS_to_ET).xlsx'
config_xlsx_full_path = os.environ.get(
    "ET_CONFIG_XLSX",
    str(Path(__file__).resolve().parent.parent / "config" / "et_100_det_inputs_1h.xlsx"),
)
runtime_config_xlsx_full_path = build_runtime_config_xlsx(
    source_path=config_xlsx_full_path,
    batch_dir=BATCH_DIR,
)

config_manager = ConfigurationManager(filepath=runtime_config_xlsx_full_path)

# ---------------------------------------------------------------------------
# 导星探测器参数覆盖：以主探测器 xlsx 为基础，仅在脚本层改写差异项。
# ---------------------------------------------------------------------------
config_manager.parameters["Simulation Name"] = "ET_guide_detector"
config_manager.parameters["Pixel Scale"] = GUIDE_PIXEL_SCALE_ARCSEC_PER_PIX * (
    u.arcsec / u.pix
)
config_manager.parameters["Pixel Width"] = GUIDE_PIXEL_WIDTH_UM * u.um
config_manager.parameters["Exposure Duration"] = GUIDE_EXPOSURE_S * u.s
config_manager.parameters["Readout Duration"] = GUIDE_READOUT_S * u.s
config_manager.parameters["Observing Duration"] = GUIDE_OBSERVING_DURATION_S * u.s
config_manager.parameters["Readout Noise"] = GUIDE_READOUT_NOISE_ADU_PER_PIX * (
    u.adu / u.pix
)
config_manager.parameters["Background Flux"] = GUIDE_BACKGROUND_FLUX_E_PER_S_PER_PIX * (
    u.electron / u.s / u.pix
)
config_manager.parameters["Scattered Light"] = (
    GUIDE_SCATTERED_LIGHT_E_PER_S_PER_PIX * u.electron / u.s / u.pix
)
config_manager.parameters["Dark Current"] = GUIDE_DARK_CURRENT_E_PER_S_PER_PIX * (
    u.electron / u.s / u.pix
)
config_manager.parameters["Inter-PRV (RMS)"] = GUIDE_INTER_PRV_RMS_PERCENT * u.percent

FRAME_TIMING = resolve_detector_frame_timing(config_manager.parameters)
RAW_FRAME_INTEGRATION_S = float(FRAME_TIMING["raw_frame_integration_s"])
RAW_FRAME_SAMPLING_INTERVAL_S = float(FRAME_TIMING["raw_frame_sampling_interval_s"])
DETECTOR_TYPE = str(FRAME_TIMING["detector_type"])

if PROFILE_TARGET_FRAMES is not None and PROFILE_TARGET_FRAMES > 0:
    config_manager.parameters["Observing Duration"] = (
        PROFILE_TARGET_FRAMES * RAW_FRAME_SAMPLING_INTERVAL_S
    ) * u.s

if not ACTIVE_EFFECT_PROFILE["uses_jitter_integrated_psf"]:
    config_manager.parameters["Use Jitter-Integrated PSF"] = False

# ---------------------------------------------------------------------------
# 关键自检：确认 xlsx 的 observing duration / cadence / 帧数
# ---------------------------------------------------------------------------
# 你在终端里看到的进度条总数（例如 x/360、x/10000）通常来自后面
# `for start in trange(0, n_coadd_frames, batch_size)` 的 `n_coadd_frames`。
# 而 `n_coadd_frames` 本质上由 xlsx 的 Observing Duration 和 cadence 决定。
_obs_s = config_manager.parameters["Observing Duration"].to(u.s).value
_exp_s = config_manager.parameters["Exposure Duration"].to(u.s).value
_rd_s = config_manager.parameters["Readout Duration"].to(u.s).value
_mult = float(config_manager.parameters["Simulation Cadence Mult"])
_frame_interval_s = RAW_FRAME_SAMPLING_INTERVAL_S
_derived_n_frames = (
    int(np.floor(_obs_s / _frame_interval_s)) if _frame_interval_s > 0 else 0
)
print(
    f"[Config check] xlsx={config_xlsx_full_path}  "
    f"obs={_obs_s:.3f}s  exp={_exp_s:.3f}s  rd={_rd_s:.3f}s  mult={_mult:g}  "
    f"detector={DETECTOR_TYPE}  integration={RAW_FRAME_INTEGRATION_S:.6f}s  "
    f"frame_interval={_frame_interval_s:.6f}s -> n_frames≈{_derived_n_frames}"
)

# ============================================================================
# 单帧尺寸修改：用导星探测器视场自动换算像素边长，并覆盖配置里的 Detector Width。
# ============================================================================
# photsim6 的 detector/PSF/image 目前主要使用一个 n_pixels（正方形）来定义单帧尺寸。
# 因此我们只需要覆盖 "Detector Width" 即可让后续：
# - mk_real_field_stars_2 查询范围（px_cnt）
# - Detector 构造
# - PSF warp
# - 生成图像数组
# 都同步变成裁剪后的大小。
pixel_scale = config_manager.parameters["Pixel Scale"].to(u.arcsec / u.pix)

# 角尺度换算：Npix = (FOV[deg] * 3600 [arcsec/deg]) / (pixel_scale [arcsec/pix])
fov_arcsec = FOV_DEG * 3600.0
n_pixels_fov = int(np.ceil(fov_arcsec / pixel_scale.value))

# 取奇数像素边长：让几何中心落在像素中心附近，便于后续中心/质心处理。
if (n_pixels_fov % 2) == 0:
    n_pixels_fov += 1

print(
    f"[Frame size] Using FOV={FOV_DEG:.2f} deg and pixel_scale={pixel_scale.value:.4f} arcsec/pix -> "
    f"Detector Width = {n_pixels_fov} pix (forced odd)."
)

config_manager.parameters["Detector Width"] = n_pixels_fov * u.pix

# 补充写入本批次 metadata（可用于你后续做“同一套 motion、不同天区中心”的对照检查）
try:
    with open(meta_path, "r", encoding="utf-8") as f:
        _meta = json.load(f)
except Exception:
    _meta = {}

_meta.update(
    {
        "instrument": "guide_detector",
        "config_xlsx_full_path": config_xlsx_full_path,
        "effect_profile": ACTIVE_EFFECT_PROFILE["name"],
        "effect_profile_requested": requested_effect_profile,
        "effect_profile_target_frames": (
            None if PROFILE_TARGET_FRAMES is None else int(PROFILE_TARGET_FRAMES)
        ),
        "effect_profile_variant_settings": dict(
            ACTIVE_EFFECT_PROFILE["variant_settings"]
        ),
        "effect_profile_components": {
            "pointing_drift": bool(ACTIVE_EFFECT_PROFILE["include_pointing_drift"]),
            "jitter": bool(
                ACTIVE_EFFECT_PROFILE["variant_settings"].get("enable_jitter", True)
            ),
            "dva": bool(ACTIVE_EFFECT_PROFILE["include_dva"]),
            "thermal": bool(ACTIVE_EFFECT_PROFILE["include_thermal"]),
            "momentum_dump": bool(ACTIVE_EFFECT_PROFILE["include_momentum_dump"]),
            "psf_breathing": bool(
                ACTIVE_EFFECT_PROFILE["variant_settings"].get(
                    "enable_psf_breathing", True
                )
            ),
            "inter_pixel_response": bool(
                ACTIVE_EFFECT_PROFILE["variant_settings"].get(
                    "enable_inter_pixel_response", True
                )
            ),
            "intra_pixel_response": bool(
                ACTIVE_EFFECT_PROFILE["variant_settings"].get(
                    "enable_intra_pixel_response", True
                )
            ),
            "pixel_phase_response": bool(
                ACTIVE_EFFECT_PROFILE["variant_settings"].get(
                    "enable_pixel_phase_response", True
                )
            ),
        },
        "psf_field_angle_deg": float(GUIDE_PSF_FIELD_ANGLE_DEG),
        "psf_field_id": int(GUIDE_PSF_FIELD_ID),
        "fov_deg": float(FOV_DEG),
        "pixel_scale_arcsec_per_pix": float(pixel_scale.value),
        "detector_width_pix": int(n_pixels_fov),
        "detector_height_pix": int(n_pixels_fov),
        "pixel_width_um": float(GUIDE_PIXEL_WIDTH_UM),
        "exposure_duration_s": float(
            config_manager.parameters["Exposure Duration"].to(u.s).value
        ),
        "readout_duration_s": float(
            config_manager.parameters["Readout Duration"].to(u.s).value
        ),
        "observing_duration_s": float(
            config_manager.parameters["Observing Duration"].to(u.s).value
        ),
        "background_surface_brightness_mag_per_arcsec2": float(
            GUIDE_BACKGROUND_SURFACE_BRIGHTNESS_MAG_PER_ARCSEC2
        ),
        "background_flux_e_per_s_per_pix": float(
            config_manager.parameters["Background Flux"]
            .to(u.electron / u.s / u.pix)
            .value
        ),
        "dark_current_e_per_s_per_pix": float(
            config_manager.parameters["Dark Current"].to(u.electron / u.s / u.pix).value
        ),
        "scattered_light_e_per_s_per_pix": float(
            config_manager.parameters["Scattered Light"]
            .to(u.electron / u.s / u.pix)
            .value
        ),
        "dark_plus_scattered_e_per_s_per_pix": float(
            config_manager.parameters["Dark Current"].to(u.electron / u.s / u.pix).value
            + config_manager.parameters["Scattered Light"]
            .to(u.electron / u.s / u.pix)
            .value
        ),
        "readout_noise_adu_per_pix": float(
            config_manager.parameters["Readout Noise"].to(u.adu / u.pix).value
        ),
        "readout_noise_e_per_pix": float(
            config_manager.parameters["Readout Noise"].to(u.adu / u.pix).value
        ),
        "gain_adu_per_electron": 1.0,
        "inter_prv_rms_percent": float(
            config_manager.parameters["Inter-PRV (RMS)"].to(u.percent).value
        ),
        "intra_prv_rms_percent": float(
            config_manager.parameters["Intra-PRV (RMS)"].to(u.percent).value
        ),
        "gaia_gmag_lim": float(GAIA_GMAG_LIM),
        "apply_static_field_offset": bool(APPLY_STATIC_FIELD_OFFSET),
        "requested_field_offset_x_pix": (
            None
            if STATIC_FIELD_OFFSET_X_PIX is None
            else float(STATIC_FIELD_OFFSET_X_PIX)
        ),
        "requested_field_offset_y_pix": (
            None
            if STATIC_FIELD_OFFSET_Y_PIX is None
            else float(STATIC_FIELD_OFFSET_Y_PIX)
        ),
        "stream_save_frames": bool(STREAM_SAVE_FRAMES),
        "frame_batch_size": int(FRAME_BATCH_SIZE),
        "frame_truth_schema_version": 2,
        "frame_truth_coordinate_convention": (
            "centered_pix uses detector-centered zero point with +y downward; "
            "image_pix uses zero-based detector array coordinates"
        ),
        "frame_truth_absolute_coordinate_fields": [
            "truth_abs_x_image_pix",
            "truth_abs_y_image_pix",
            "truth_abs_x_detector_pix",
            "truth_abs_y_detector_pix",
        ],
        "frame_truth_source": "photsim6ft_frame_truth_builder",
        "field_polar_angle_rad": float(FIELD_POLAR_ANGLE_RAD),
        "jitter_truth_mode": "shared_per_frame_mean",
    }
)

with open(meta_path, "w", encoding="utf-8") as f:
    json.dump(_meta, f, ensure_ascii=False, indent=2)

# config_manager.parameters['Observing Duration'] = 7 * u.day

# config_manager.parameters

variant_manager = VariantManager()  # Initialize the manager

# Add variant 0 - active effect profile
variant_manager.add_variant(
    description=ACTIVE_EFFECT_PROFILE["description"],
    optimal_aperture=config_manager.parameters["Optimal Aperture Algorithim"],
    **ACTIVE_EFFECT_PROFILE["variant_settings"],
)

"""Example extra variant"""
# variant_manager.add_variant(
#     description='1',
#     optimal_aperture=config_manager.parameters['Optimal Aperture Algorithim'],
#     enable_stellar_photon_noise=False,
# )

# accepted_settings = {
#     'description': 'None',
#     'enable_stellar_photon_noise': True,
#     'enable_background_light': True,
#     'enable_scattered_light': True,
#     'enable_dark_current': True,
#     'enable_readout_noise': True,
#     'enable_gain': True,
#     'enable_target_star': True,
#     'enable_background_stars': True,
#     'enable_jitter': True,
#     'enable_dva_drift': True,
#     'enable_pointing_drift': True,
#     'enable_psf_breathing': True,
#     'enable_inter_pixel_response': True,
#     'enable_intra_pixel_response': True,
#     'enable_pixel_phase_response': True,
#     'enable_flat_field_correction': False,
#     'use_oa_from_variant': 0,
#     'optimal_aperture': 'Kepler',
#     'enable_coadding': True,
# }

# variant_manager.add_variant(
#     description='1',
#     optimal_aperture=config_manager.parameters['Optimal Aperture Algorithim'],
#     enable_dva_drift=False,
#     enable_pointing_drift=False,
#     use_oa_from_variant=1,
# )

# variant_manager.add_variant(
#     description='2',
#     optimal_aperture=config_manager.parameters['Optimal Aperture Algorithim'],
#     enable_jitter=False,
#     use_oa_from_variant=2,
# )

# variant_manager.add_variant(
#     description='3',
#     optimal_aperture=config_manager.parameters['Optimal Aperture Algorithim'],
#     enable_psf_breathing=False,
#     use_oa_from_variant=3,
# )

# ---------------------------------------------------------------------------------------------

print(
    "--------------------------------------------------\nUser-defined variant catalog\n\n"
)
# Display the user-defined variant catalog
# display(variant_manager.variants_dataframe)
print(
    "\n--------------------------------------------------\nUser-defined + optimal aperture required variant catalog\n\n"
)
# Display the user-defined + optimal aperture required variants catalog
# display(variant_manager.all_variants_dataframe)

del variant_manager.all_variants[1]
del variant_manager.all_variants[2]

variant_manager._update_all_variants_dataframe()

print(
    "--------------------------------------------------\nUser-defined variant catalog\n\n"
)
# Display the user-defined variant catalog
# display(variant_manager.variants_dataframe)
print(
    "\n--------------------------------------------------\nUser-defined + optimal aperture required variant catalog\n\n"
)
# Display the user-defined + optimal aperture required variants catalog
# display(variant_manager.all_variants_dataframe)

# Load and parse motion power spectrum data into a Dataframe (DF)
power_spec_config = load_power_spectrum_config(config_xlsx_full_path)

# Build the power spectrum dictionary from the DF
comps = create_power_spec_components(power_spec_config)

# display(power_spec_config)  # Show DF

comps  # Show dict

# Build the motion power spectrum
jit_spec_mask, jitter_freq_mag, drift_spec_mask, drift_freq_mag = gen_psd_motion(
    comps, plot=False
)

et_freqs, et_amps, jit_freq_range = gen_power_spec(comps)

mask2 = (et_freqs >= 1 / (5 * u.hr.to(u.s))) & (et_freqs <= 1 / (30 * u.min.to(u.s)))
et_amps[mask2] = 0.00014079 * 1.8


# -----------------------------------------------------------------------------
# Drift/Jitter 分界（脚本层实现）：split_hz = 1 / raw-frame integration time
# -----------------------------------------------------------------------------
# 这里统一使用：
# - ET PSD：只保留 <= split_hz 的低频部分，作为“帧间漂移/慢变化”
# - TESS PSD：只保留 >  split_hz 的高频部分，作为“曝光内抖动/快变化”
# 这样可以避免把本应留下帧间质心差异的运动误当作 jitter 进入 PSF 卷积。
raw_frame_integration_s = RAW_FRAME_INTEGRATION_S
raw_frame_sampling_interval_s = RAW_FRAME_SAMPLING_INTERVAL_S

if USE_CADENCE_SPLIT:
    split_hz = 1.0 / float(raw_frame_integration_s)
else:
    # 兼容旧逻辑（历史脚本用 0.1Hz）：不建议用于 1.6s cadence 的导星评估。
    split_hz = 0.1

print(
    f"[Motion split] detector={DETECTOR_TYPE}  integration={raw_frame_integration_s:.3f}s  "
    f"frame_interval={raw_frame_sampling_interval_s:.3f}s -> split_hz={split_hz:.6f} Hz "
    f"(drift<=split, jitter>split)."
)

if SPLIT_HZ_OVERRIDE is not None:
    split_hz = float(SPLIT_HZ_OVERRIDE)
    print(
        f"[Motion split] SPLIT_HZ_OVERRIDE is set -> split_hz={split_hz:.6f} Hz "
        f"(drift<=split, jitter>split)."
    )

# 只保留 ET 的低频部分（<= split_hz），高频部分由 TESS jitter 负责。
et_amps = et_amps.copy()
et_amps[et_freqs > split_hz] = 0.0

et_amp_spec = [et_freqs, et_amps]

pixel_scale = config_manager.parameters["Pixel Scale"]

plot_psd(comps, et_freqs, et_amps)
plt.plot(et_freqs, et_amps * pixel_scale.value, ":c", label="All")
plt.legend()

len(jit_spec_mask), jit_spec_mask.sum(), len(drift_spec_mask), drift_spec_mask.sum()

tess_psd = load_pickle("/home/cxgao/ET/Photosim6/data/tess_data/tess_xyz_psd.pkl")

"""


IS SET BELOW TO MAKE 2X TESS JITTER !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!


"""

tess_jitter_mult = 2  # Multiply the amplitude of TESS jitter

# ============================================================================
# JI-PSF jitter 序列生成（只用于“曝光内模糊”）：只取 > split_hz 的高频部分
# ============================================================================
# 重要：photsim6 的 jitter-integrated PSF 是“把 PSF 在一条 jitter 轨迹上平移多帧，再求平均”。
# 因此 xy_jitter_pix 的每个条目应该代表“一次曝光内”的快抖动轨迹。
# 我们把 <= split_hz 的慢变化都剥离到 drift（每帧中心偏移）里，避免把慢漂移错误地卷进 PSF。


def _angles_to_xy_pix(
    theta_x_arcsec,
    theta_y_arcsec,
    theta_z_arcsec,
    field_angle_deg,
    x_axis_angle_deg,
    plate_scale,
):
    """把姿态角(arcsec)转换为焦平面上的 x/y 像素位移。

    这里复用 photsim6ft 的思路：
    - 用一个代表全幅焦平面坐标系的位置 (x_pix, y_pix) 来施加旋转
    - displacement = old - new

    注意：即使我们只仿真裁剪图（n_pixels 变小），旋转导致的位移仍由“真实 field_angle”决定，
    与你裁剪多少像素无关。所以这里仍然用 23.5°/9000pix 的全幅映射来取参考点。
    """
    from photsim6.data_generators import PixelSpaceSimulator

    x_center = 0.0
    y_center = 0.0
    et_pss = PixelSpaceSimulator(
        plate_scale=plate_scale, x_center=x_center, y_center=y_center
    )

    fa_theta_rad = x_axis_angle_deg / 180.0 * np.pi
    x_fa = np.cos(fa_theta_rad) * field_angle_deg
    y_fa = np.sin(fa_theta_rad) * field_angle_deg

    x_pix = np.interp(x_fa, [0, 23.5 / 2], [0, 9000])
    y_pix = np.interp(y_fa, [0, 23.5 / 2], [0, 9000])

    x_pix_new, y_pix_new = et_pss.apply_spacecraft_rotations(
        theta_x_arcsec * u.arcsec,
        theta_y_arcsec * u.arcsec,
        theta_z_arcsec * u.arcsec,
        x_pix,
        y_pix,
    )

    dx_pix = x_pix - x_pix_new
    dy_pix = y_pix - y_pix_new
    return dx_pix, dy_pix


def generate_tess_centroid_xy_motion_split(
    time,
    tess_psd,
    split_hz,
    field_angle_deg=10.0,
    x_axis_angle_deg=45.0,
    tess_mult=1.0,
    highpass=True,
    plot=False,
):
    """从 TESS PSD 生成 x/y 像素位移，并按 split_hz 做高/低频分割。

    - highpass=True  -> 只取 > split_hz 的高频部分（jitter，供 JI-PSF 使用）
    - highpass=False -> 只取 <= split_hz 的低频部分（drift，供每帧中心偏移使用）

    这里显式调用 `generate_tess_jitter()` 而不是用包内 `generate_tess_centroid_jitter()`：
    因为后者内部把 x/y 的 frequency_min 硬编码成 0.1Hz，不适配你要的 0.625Hz 分界。
    """
    from photsim6.data_generators import generate_tess_jitter

    def _resample_to_time_grid(values, src_time_s, dst_time_s):
        """把 generate_tess_jitter() 的输出重采样到给定的 time 网格上。

        背景：generate_tess_jitter 内部用 `N_time = int(duration * f_samp) + 1`。
        当 duration*f_samp 因浮点误差略小于整数时，会少 1 个采样点（例如 9999 而不是 10000），
        导致后续与 `time` 的长度不一致并触发 broadcast 错误。
        """
        values = np.asarray(values, dtype=float)
        dst_time_s = np.asarray(dst_time_s, dtype=float)
        if values.shape[0] == dst_time_s.shape[0]:
            return values

        src_time_s = np.asarray(src_time_s, dtype=float)
        if src_time_s.shape[0] != values.shape[0]:
            src_time_s = np.linspace(dst_time_s[0], dst_time_s[-1], num=values.shape[0])
        if src_time_s[0] > src_time_s[-1]:
            src_time_s = src_time_s[::-1]
            values = values[::-1]

        return np.interp(dst_time_s, src_time_s, values)

    ztime_s = (time - time[0]).to(u.s).value
    if ztime_s.size >= 2:
        dt = float(ztime_s[1] - ztime_s[0])
    else:
        dt = float(raw_frame_sampling_interval_s)
    f_samp = 1.0 / dt
    # 用 (t[-1] + dt) 而不是 t[-1]，避免 duration*f_samp 因浮点误差落到整数的左侧。
    duration = float(ztime_s[-1] + dt)

    if highpass:
        fmin = float(split_hz)
        fmax = np.inf
    else:
        fmin = 0.0
        fmax = float(split_hz)

    x_tess_freqs, x_tess_psd = tess_psd["x"]
    y_tess_freqs, y_tess_psd = tess_psd["y"]
    z_tess_freqs, z_tess_psd = tess_psd["z"]

    theta_x_as, t_x = generate_tess_jitter(
        x_tess_freqs.value,
        x_tess_psd.value,
        frequency_min=fmin,
        frequency_max=fmax,
        duration=duration,
        base_f_samp=f_samp,
        supersample_factor=1,
    )
    theta_y_as, t_y = generate_tess_jitter(
        y_tess_freqs.value,
        y_tess_psd.value,
        frequency_min=fmin,
        frequency_max=fmax,
        duration=duration,
        base_f_samp=f_samp,
        supersample_factor=1,
    )
    theta_z_as, t_z = generate_tess_jitter(
        z_tess_freqs.value,
        z_tess_psd.value,
        frequency_min=fmin,
        frequency_max=fmax,
        duration=duration,
        base_f_samp=f_samp,
        supersample_factor=1,
    )

    # 兜底：确保输出与 time 的采样点数严格一致
    theta_x_as = _resample_to_time_grid(theta_x_as, t_x, ztime_s)
    theta_y_as = _resample_to_time_grid(theta_y_as, t_y, ztime_s)
    theta_z_as = _resample_to_time_grid(theta_z_as, t_z, ztime_s)

    # 与旧脚本一致：仅放大 x/y 的抖动幅度（tess_mult），z 轴不放大。
    theta_x_as = theta_x_as * tess_mult
    theta_y_as = theta_y_as * tess_mult

    if plot:
        plt.figure(figsize=(11, 4))
        plt.plot(ztime_s, theta_x_as, ".-", label="theta_x [arcsec]")
        plt.plot(ztime_s, theta_y_as, ".-", label="theta_y [arcsec]")
        plt.plot(ztime_s, theta_z_as, ".-", label="theta_z [arcsec]")
        plt.title(
            f"TESS {'high' if highpass else 'low'}-freq motion "
            f"(f in [{fmin}, {fmax}])"
        )
        plt.legend()
        plt.tight_layout()
        plt.show()

    dx_pix, dy_pix = _angles_to_xy_pix(
        theta_x_arcsec=theta_x_as,
        theta_y_arcsec=theta_y_as,
        theta_z_arcsec=theta_z_as,
        field_angle_deg=field_angle_deg,
        x_axis_angle_deg=x_axis_angle_deg,
        plate_scale=pixel_scale,
    )

    return dx_pix, dy_pix


def generate_et_xy_motion_lowfreq(time, et_amp_spec, split_hz):
    """从 ET PSD 生成低频(<=split_hz)的 x/y 像素位移，用作 drift。"""
    ztime_s = (time - time[0]).to(u.s).value
    et_f, et_amp = et_amp_spec

    # 低频漂移：<= split_hz
    mask = (et_f <= split_hz) & (et_amp > 0)
    if mask.sum() == 0:
        return np.zeros_like(ztime_s, dtype=float), np.zeros_like(ztime_s, dtype=float)

    xy_phases = [np.random.uniform(0, 2 * np.pi, mask.sum()) for _ in range(2)]
    x_pix = psd_to_motion(et_f[mask], et_amp[mask], xy_phases[0], ztime_s)
    y_pix = psd_to_motion(et_f[mask], et_amp[mask], xy_phases[1], ztime_s)
    return x_pix, y_pix


def generate_jitter_xy_highfreq_for_exposure(
    time, et_amp_spec, tess_psd, split_hz, tess_mult=1.0, plot=False
):
    """生成一次曝光内的高频 jitter 轨迹（>split_hz），供 JI-PSF 模糊使用。

    这里选择：
    - ET 高频部分已被上面 et_amps[et_freqs>split]=0 剔除（避免与 TESS jitter 叠加）。
    - jitter 统一来自 TESS PSD 的高频部分 (> split_hz)。
    """
    x_jit_tess_pix, y_jit_tess_pix = generate_tess_centroid_xy_motion_split(
        time=time,
        tess_psd=tess_psd,
        split_hz=split_hz,
        field_angle_deg=10.0,
        x_axis_angle_deg=45.0,
        tess_mult=tess_mult,
        highpass=True,
        plot=plot,
    )
    return x_jit_tess_pix, y_jit_tess_pix


resume = True

use_jitter_integrated_psf = bool(ACTIVE_EFFECT_PROFILE["uses_jitter_integrated_psf"])

# Get key params for jitter generation and the filename
n_jitter_frames_per_model = int(config_manager.parameters["N Jitter Frames Per Model"])
n_jit_psf = config_manager.parameters["N Jitter-Integrated PSF Models"]

jitter_mag_str = str(np.around(comps["Jitter: All"]["mag"], 7))
jitter_frq_str = str(comps["Jitter: All"]["freqs"]).replace(" ", "")

# Define filename and save/load directory
# jit_fn = f'xy_jitter_pix_N({n_jit_psf})_AFJ({jitter_mag_str}_{jitter_frq_str})_exp{raw_frame_integration_s:.2f}s.npy'

# 文件名里显式写入 split_hz，避免你切换 cadence 后误读旧缓存。
jit_fn = (
    f"xy_{tess_jitter_mult}X_jitter_pix_N({n_jit_psf})_"
    f"AFJ(Rescaled_TESS_PSD)_{DETECTOR_TYPE}_"
    f"exp{raw_frame_integration_s:.6f}s_"
    f"nsub{n_jitter_frames_per_model}_"
    f"split{split_hz:.6f}Hz.npy"
)

jit_fp = opj(CACHE_DIR, "jitter", "et", "JI-PSF_data", jit_fn)

os.makedirs(os.path.dirname(jit_fp), exist_ok=True)

# -------------------------------------------------------------------

if use_jitter_integrated_psf:
    if os.path.isfile(jit_fp) and resume:
        print("Loaded jitter")
        xy_jitter_pix = np.load(jit_fp)
    else:
        xy_jitter_pix = []

        for i in trange(
            config_manager.parameters["N Jitter-Integrated PSF Models"]
        ):  # For each JI-PSF

            et_time = (
                np.linspace(
                    0.0,
                    raw_frame_integration_s,
                    num=n_jitter_frames_per_model,
                    endpoint=False,
                    dtype=np.float64,
                )
                * u.s
            )

            # 只生成 > split_hz 的高频抖动：用于“曝光内 PSF 模糊”
            x_jit_pix, y_jit_pix = generate_jitter_xy_highfreq_for_exposure(
                time=et_time,
                et_amp_spec=et_amp_spec,
                tess_psd=tess_psd,
                split_hz=split_hz,
                tess_mult=tess_jitter_mult,
                plot=False,
            )

            xy_jitter_pix.append([x_jit_pix, y_jit_pix])

        print("Saved jitter")
        np.save(jit_fp, xy_jitter_pix)
else:
    xy_jitter_pix = np.zeros((1, 2, 1), dtype=np.float32)
    print(
        '[Effect profile] Skip jitter cache and force "Use Jitter-Integrated PSF" to FALSE.'
    )

jit_fp


def generate_tess_roll_xy_drift(
    time, tess_psd, split_hz, field_angle=10.0, x_axis_angle=45
):

    from photsim6.data_generators import PixelSpaceSimulator

    x_center = 0.0
    y_center = 0.0

    et_pss = PixelSpaceSimulator(
        plate_scale=pixel_scale,
        x_center=x_center,
        y_center=y_center,
    )
    # -----------------------------------------------------------------------------------
    # TESS 低频(<=split_hz) roll motion -> x, y motion（作为 drift 的一部分）

    def _resample_to_time_grid(values, src_time_s, dst_time_s):
        values = np.asarray(values, dtype=float)
        dst_time_s = np.asarray(dst_time_s, dtype=float)
        if values.shape[0] == dst_time_s.shape[0]:
            return values

        src_time_s = np.asarray(src_time_s, dtype=float)
        if src_time_s.shape[0] != values.shape[0]:
            src_time_s = np.linspace(dst_time_s[0], dst_time_s[-1], num=values.shape[0])
        if src_time_s[0] > src_time_s[-1]:
            src_time_s = src_time_s[::-1]
            values = values[::-1]
        return np.interp(dst_time_s, src_time_s, values)

    ztime_s = (time - time[0]).to(u.s).value
    if ztime_s.size >= 2:
        dt = float(ztime_s[1] - ztime_s[0])
    else:
        dt = float(raw_frame_sampling_interval_s)
    f_samp = 1.0 / dt
    duration = float(ztime_s[-1] + dt)

    from photsim6.data_generators import generate_tess_jitter

    z_tess_freqs, z_tess_psd = tess_psd["z"]
    theta_z_arcsec, t_z = generate_tess_jitter(
        z_tess_freqs.value,
        z_tess_psd.value,
        frequency_min=0.0,
        frequency_max=float(split_hz),
        duration=duration,
        base_f_samp=f_samp,
        supersample_factor=1,
    )

    theta_z_arcsec = _resample_to_time_grid(theta_z_arcsec, t_z, ztime_s)

    fa_theta_rad = x_axis_angle / 180 * np.pi
    x_fa = np.cos(fa_theta_rad) * field_angle
    y_fa = np.sin(fa_theta_rad) * field_angle

    x_pix = np.interp(x_fa, [0, 23.5 / 2], [0, 9000])
    y_pix = np.interp(y_fa, [0, 23.5 / 2], [0, 9000])

    x_pix_new, y_pix_new = et_pss.apply_spacecraft_rotations(
        0 * u.arcsec,
        0 * u.arcsec,
        theta_z_arcsec * u.arcsec,
        x_pix,
        y_pix,
    )

    x_roll_pix = x_pix - x_pix_new
    y_roll_pix = y_pix - y_pix_new

    return x_roll_pix, y_roll_pix, theta_z_arcsec


time_example = np.arange(0, 2 * u.day.to(u.s), 10) * u.s

# theta_z = get_roll_angles(time_example)
x_drift_tess_roll_pix, y_drift_tess_roll_pix, z_drift_theta_as = (
    generate_tess_roll_xy_drift(time_example, tess_psd, split_hz=split_hz)
)

print(
    f"Generated {len(z_drift_theta_as)} roll angles for times {time_example[0]} to {time_example[-1]} s"
)
print("First few roll angles [arcsec]:", z_drift_theta_as[:5])

from photsim6.plot import centroid_motion_and_component_plot_with_r95

r_stat_circles = [95]
r_stat_colors = None

fig = centroid_motion_and_component_plot_with_r95(
    t=time_example.to(u.s),
    x=x_drift_tess_roll_pix * pixel_scale.value,
    y=y_drift_tess_roll_pix * pixel_scale.value,
    title="",
    name="Drift",
    xy_unit="arcsec",
    time_unit="day",
    figsize=(16, 6),
    r_stat_circles=r_stat_circles,
    r_stat_colors=r_stat_colors,
)

# Define dynamic param configs (DPC)

# ------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------
# Build the main dict that holds all DPC
dynamic_param_config = {}
dynamic_param_config["motion"] = []

# ---------------------------------
# Momentum dump of spacecraft reaction wheels

if ACTIVE_EFFECT_PROFILE["include_momentum_dump"] and (
    config_manager.parameters["Momentum Dump Model"].lower() != "none"
):  # Was MD requested?

    # Get user requested MD settings
    jump_cadence = config_manager.parameters["Momentum Dump Cycle"]
    jump_length = config_manager.parameters["Momentum Dump R(68%)"]
    jump_circ_rad = config_manager.parameters["Momentum Dump R(68%)"]

    obs_duration = config_manager.parameters["Observing Duration"]
    if obs_duration.to(u.s).value < jump_cadence.to(u.s).value:
        print(
            "[Momentum dump] skip: observing duration "
            f"{obs_duration.to(u.s).value:.3f}s < cycle {jump_cadence.to(u.s).value:.3f}s."
        )
    else:
        # Create the MD DPC item
        dynamic_param_config["motion"] += [  # Its motion, so add to the motion catagory
            dict(
                component_name="momentum_dump_jumps",
                model_name=config_manager.parameters["Momentum Dump Model"],
                model_params=dict(
                    period_step=jump_cadence,
                    circle_radius=jump_circ_rad.to(u.arcsec).value / pixel_scale.value,
                    r_step_avg=jump_length.to(u.arcsec).value / pixel_scale.value,
                    stay_inside=True,
                    random_r_step=True,
                    step_func=True,
                ),
            ),
        ]

# ---------------------------------
# DVA motion

dva_model_data = load_pickle(
    opj(BASE_DATA_DIR, "DVA", "et", "ET_DVA_effect_models_slim_v231117.pkl")
)  # Load DVA model (created from JWST code)

if ACTIVE_EFFECT_PROFILE["include_dva"]:
    dynamic_param_config["motion"] += [
        dict(
            component_name="dva_drift",
            model_name="dva_model",  # Built-in DVA model
            model_params=dict(
                dva_model=dva_model_data,
                psf_field_angle=12.0,  # Field angle (FA) of 12 deg is where mots stars will be observed in ET
                pixel_scale=pixel_scale.value,
                t0=0.0,
            ),
        ),
    ]

# ---------------------------------
# Thermal drift due to thermal defocus
"""
Based on TESS lens temperatures and comverted to ET motion using a
ET lens temp to centorid motion model provided by ET engineers.
"""

if ACTIVE_EFFECT_PROFILE["include_thermal"]:
    dynamic_param_config["motion"] += [
        dict(
            component_name="thermal_drift",
            model_name="et_tess_thermal_drift_model",
            model_params=dict(
                field_angle=10.0,  # Cant change until more data from ET engineers..
            ),
        ),
    ]

# ---------------------------------
# Other drifts (from power spectrum)
"""
This eventually needs to be better implemented. Currently, it uses a custom model
which is defined in the function below (uses global variables! Ah!).
"""


def generate_fft_hf_drift(time):

    ztime_s = (time - time[0]).to(u.s).value

    def _ensure_len_on_ztime(arr):
        arr = np.asarray(arr, dtype=float)
        if arr.ndim == 0:
            return np.full(ztime_s.shape, float(arr), dtype=float)
        if arr.shape[0] == ztime_s.shape[0]:
            return arr
        src_t = np.linspace(ztime_s[0], ztime_s[-1], num=arr.shape[0])
        return np.interp(ztime_s, src_t, arr)

    # =====================================================================
    # 选项A（脚本层分界）：生成每帧中心漂移 drift（<= split_hz）
    # =====================================================================
    # 这里返回的 x/y 会被 image.py 叠加到每帧 PSF 中心位置上（frame-to-frame motion）。
    # 与之对应的“曝光内高频 jitter”已经通过 xy_jitter_pix 进入 JI-PSF 模糊。

    # 1) ET 低频 drift（<= split_hz）
    x_drift_et_pix, y_drift_et_pix = generate_et_xy_motion_lowfreq(
        time=time, et_amp_spec=et_amp_spec, split_hz=split_hz
    )
    x_drift_et_pix = _ensure_len_on_ztime(x_drift_et_pix)
    y_drift_et_pix = _ensure_len_on_ztime(y_drift_et_pix)

    # 2) TESS 低频 roll drift（<= split_hz）
    x_drift_tess_roll_pix, y_drift_tess_roll_pix, _ = generate_tess_roll_xy_drift(
        time=time,
        tess_psd=tess_psd,
        split_hz=split_hz,
        field_angle=10.0,
        x_axis_angle=45,
    )
    x_drift_tess_roll_pix = _ensure_len_on_ztime(x_drift_tess_roll_pix)
    y_drift_tess_roll_pix = _ensure_len_on_ztime(y_drift_tess_roll_pix)

    # 3) 合并得到总 drift（不包含 DVA/thermal/MD 等，它们在其他 motion 组件里单独加）
    x_drift_pix = x_drift_et_pix + x_drift_tess_roll_pix
    y_drift_pix = y_drift_et_pix + y_drift_tess_roll_pix
    return x_drift_pix, y_drift_pix


if ACTIVE_EFFECT_PROFILE["include_pointing_drift"]:
    dynamic_param_config["motion"] += [
        dict(
            component_name="fft_hf_drift",
            model_name="user_input_xy_function",  # Tells the generator to use a custom model function
            model_params=dict(
                func=generate_fft_hf_drift,
            ),
        ),
    ]

# ------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------
# PSF breathing
"""
This eventually needs to be better implemented. Currently, it uses a custom model
which is defined in the function imported below. This function is similar to the
thermal_drift DP in that it uses TESS lens temperatures and an ET temp to PSF breathing
model (from engineers) to generate PSF breathing in the simulation.
"""
from photsim6.data_generators import et_time_to_breathing

dynamic_param_config["psf_scale"] = []

dynamic_param_config["psf_scale"] += [
    # dict(
    #     component_name='psf_scale',
    #     model_name='user_input_r_function',
    #     model_params=dict(
    #         func=et_time_to_breathing, # ET thermal PSF breathing derived from TESS temperatures and generated using ET field angle of ONLY 10 deg!
    #     ),
    # ),
    dict(
        component_name="psf_scale",
        model_name="sine_1d",
        model_params=dict(
            amp=3.0 / 100,
            period=3.0 * u.day,
            offset=1.0,
        ),
    ),
]

# ------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------

# Display the motion-only DPC dict, eventually provided to the simulator
dynamic_param_config["motion"], dynamic_param_config["psf_scale"]

inject_transits = False
# inject_transits = True

if inject_transits:

    n_exop = 20  # How many transits to be injected in a batch of stars.
    n_stars = 500  # How many stars will you simulate per batch? Typically simulate 500 stars per batch.

    # Build the dict that will be provided to the simulator to generate 4 earth-like transits per light curve, regardless of observing length
    transit_data = dict(
        star_ids=np.sort(
            np.random.permutation(n_stars)[:n_exop]
        ),  # In the batch, which star IDs will have transits injected?
        star_mags=np.full(
            n_exop, 13.0
        ),  # What should be the magnitude of each star with transits?
        to_img_gen=[],  # The relative flux profile for each transit. If not provided, it will be autmatically generated for you.
    )

else:
    transit_data = None

transit_data

config_manager.parameters["Telescope Count"] = 1

simulator = Simulator(
    sim_config=config_manager.parameters,
    dynamic_param_config=dynamic_param_config,
    variant_manager=variant_manager,
    n_raw_frames_per_coadd=1,
    xy_jitter_pix=xy_jitter_pix,
    compute_device=COMPUTE_DEVICE,
    float_precision=32,
    mag_range=[7, 17],
    # store_images=True,
    start_ray=False,
    transit_data=transit_data,
    ray_cluster_address=None,
    ray_namespace=None,
    mag_type="ET",
    # ray_cluster_address='ray://100.125.174.11:10001',
    # ray_namespace="persistent_actors",
)

simulator.time_manager.summary

# Confirm the readout noise for coadded frames (RN * sqrt(n_raw_frames_per_coadd))
simulator.sim_config["Sim Readout Noise"]


from photsim6.psf.model import PSFModelManager

# from photsim6.psf.utils import generate_random_field_id
# from photsim6.lc_processing import generate_variant_light_curves
from photsim6.image import ImageProcessing, generate_images_for_star

# from photsim6.aperture import build_all_variant_oa
# # from photsim6.analysis import measure_centroids
from photsim6.field import Stars, generate_field_stars

# from photsim6.instrumentation import Observatory

actor_config = simulator.build_actor_config()

# 进一步补充本批次 metadata：帧数/时间采样等
try:
    with open(meta_path, "r", encoding="utf-8") as f:
        _meta = json.load(f)
except Exception:
    _meta = {}

_meta.update(
    {
        "requested_compute_device": REQUESTED_COMPUTE_DEVICE,
        "compute_device": COMPUTE_DEVICE,
        "detector_type": DETECTOR_TYPE,
        "telescope_count": int(config_manager.parameters["Telescope Count"]),
        "n_coadd_frames": int(actor_config.get("n_coadd_frames")),
        "sim_frame_duration_s": float(
            simulator.time_manager.sim_frame_duration.to(u.s).value
        ),
        "sim_exposure_s": float(simulator.time_manager.sim_exposure.to(u.s).value),
        "sim_readout_s": float(simulator.time_manager.sim_readout.to(u.s).value),
        "raw_frame_integration_s": float(raw_frame_integration_s),
        "raw_frame_sampling_interval_s": float(raw_frame_sampling_interval_s),
        "motion_split_hz": float(split_hz),
        "motion_split_reference": "raw_frame_integration_s",
        "n_jitter_frames_per_model": int(n_jitter_frames_per_model),
        "use_jitter_integrated_psf": bool(use_jitter_integrated_psf),
    }
)

with open(meta_path, "w", encoding="utf-8") as f:
    json.dump(_meta, f, ensure_ascii=False, indent=2)

psf_model_manager = PSFModelManager(
    config=actor_config,
    warp_frame_batch_size=10,
    xy_jitter_pix=None,
    intialize=True,
    build_jit_int_models=use_jitter_integrated_psf,
)

psf_field_id = GUIDE_PSF_FIELD_ID  # 14 deg guide-detector PSF

psf_model = psf_model_manager.models[psf_field_id]
# star_data['id'] = star_id

# simulator.observatory = simulator.Observatory(simulator.sim_config, psf_model.compute_device, psf_model.float_precision)
simulator.observatory.store_variant_detectors(simulator.variant_manager)

# ============================================================================
# 调用 mk_real_field_stars 生成真实星场
# ============================================================================

from photsim6.field import mk_real_field_stars_2
from photsim6.field import mk_real_field_stars_et_focalplane
from astropy import units as u
import numpy as np
import pandas as pd

# 从配置中获取参数
pixel_scale = config_manager.parameters[
    "Pixel Scale"
]  # 像素尺度，例如: 4.83 arcsec/pix
n_pixels = int(
    config_manager.parameters["Detector Width"].to(u.pix).value
)  # 探测器宽度，例如: 1001

# 从 tile_centers.csv 读取第一组中心坐标
# 注意：如果文件不在当前目录，请修改路径
# tile_csv_path = 'tile_centers.csv'  # 修改为实际路径，例如: '/path/to/tile_centers.csv'

# tile_centers_df = pd.read_csv(tile_csv_path)
# i = 1
# first_tile = tile_centers_df.iloc[i]  # 第一行数据（第一组中心）

# target_ra = first_tile['center_ra_deg']   # 赤经（度）
# target_dec = first_tile['center_dec_deg'] # 赤纬（度）

# 使用脚本顶部提供的“天区中心接口”，便于你快速改 4 批中心。
target_ra = float(FIELD_CENTER_RA_DEG)  # 赤经（度）
target_dec = float(FIELD_CENTER_DEC_DEG)  # 赤纬（度）

star_data, guide_query_meta = query_guide_star_field_adaptive(
    ra=target_ra * u.deg,  # 目标星赤经（转换为 Quantity）
    dec=target_dec * u.deg,  # 目标星赤纬（转换为 Quantity）
    px_cnt=n_pixels,  # 探测器像素数
    px_scale=pixel_scale,  # 像素尺度
    apply_offset=APPLY_STATIC_FIELD_OFFSET,  # 是否添加整幅星场静态偏移
    plot=True,  # 仅在最终采用的查询深度上绘制星场图
    offset_x_pix=STATIC_FIELD_OFFSET_X_PIX,
    offset_y_pix=STATIC_FIELD_OFFSET_Y_PIX,
    base_gaia_gmag_lim=GAIA_GMAG_LIM,
    min_catalog_stars=GUIDE_MIN_CATALOG_STARS,
    gaia_gmag_step=GAIA_GMAG_LIM_STEP,
    gaia_gmag_lim_max=GAIA_GMAG_LIM_MAX,
    query_backend=GUIDE_STAR_QUERY_BACKEND,
    catalog_dir=GUIDE_GAIA_CATALOG_DIR,
    et_focalplane_data_dir=ET_FOCALPLANE_DATA_DIR,
    et_focalplane_src_dir=ET_FOCALPLANE_SRC_DIR,
    target_epoch=GUIDE_QUERY_TARGET_EPOCH,
    crop_to_simulation_frame=GUIDE_QUERY_CROP_TO_SIM_FRAME,
    crop_margin_pix=GUIDE_QUERY_CROP_MARGIN_PIX,
)
star_data = enrich_star_data_truth_columns(star_data)

try:
    with open(meta_path, "r", encoding="utf-8") as f:
        _meta = json.load(f)
except Exception:
    _meta = {}

_meta.update(
    {
        "field_offset_x_pix": float(star_data.get("field_offset_x_pix", 0.0)),
        "field_offset_y_pix": float(star_data.get("field_offset_y_pix", 0.0)),
        "guide_query_target_center_xpix": float(
            star_data.get("target_center_xpix", np.nan)
        ),
        "guide_query_target_center_ypix": float(
            star_data.get("target_center_ypix", np.nan)
        ),
        "guide_query_target_center_xpix_shifted": float(
            star_data.get("target_center_xpix_shifted", np.nan)
        ),
        "guide_query_target_center_ypix_shifted": float(
            star_data.get("target_center_ypix_shifted", np.nan)
        ),
        "guide_query_backend": str(guide_query_meta["query_backend"]),
        "guide_query_detector_id": str(guide_query_meta["query_detector_id"]),
        "guide_query_target_epoch": float(guide_query_meta["query_target_epoch"]),
        "guide_query_base_gaia_gmag_lim": float(
            guide_query_meta["query_base_gaia_gmag_lim"]
        ),
        "guide_query_final_gaia_gmag_lim": float(
            guide_query_meta["query_final_gaia_gmag_lim"]
        ),
        "guide_query_gaia_gmag_step": float(guide_query_meta["query_gaia_gmag_step"]),
        "guide_query_gaia_gmag_lim_max": float(
            guide_query_meta["query_gaia_gmag_lim_max"]
        ),
        "guide_query_crop_to_simulation_frame": bool(
            guide_query_meta["query_crop_to_simulation_frame"]
        ),
        "guide_query_crop_margin_pix": float(guide_query_meta["query_crop_margin_pix"]),
        "guide_query_history": guide_query_meta["query_history"],
        "guide_query_reached_min_star_target": bool(
            guide_query_meta["query_reached_min_star_target"]
        ),
    }
)

with open(meta_path, "w", encoding="utf-8") as f:
    json.dump(_meta, f, ensure_ascii=False, indent=2)

n_stars = len(star_data["x0"])
# Build Stars catalog
stars = Stars()
optical_eff_dpct = (
    simulator.sim_config["Optical Efficiency Ratio"].to(u.percent).value / 100
)
stars.build_catalog(
    star_data,
    simulator.time_manager.sim_exposure,
    optical_eff_dpct,
    simulator.relative_aperture_area,
    mag_type=simulator.mag_type,
)

field_id = psf_field_id  # Which PSF ID to use? Keep catalog/ray generation consistent with the selected PSF.
star_id = 0  # Bogus/doesnt matterprint(np.sum(mask), column_sum)
transit_rel_flux = None
# Assign field ID and FOV angle
star_data["field_id"] = field_id
stars.catalog["Field ID"] = field_id
stars.catalog["Star ID"] = star_id

target_star_angle = float(FIELD_POLAR_ANGLE_RAD)
star_data["fov_theta"] = target_star_angle
stars.catalog["FOV theta"] = target_star_angle
if "source_id" in star_data:
    stars.catalog["Source ID"] = np.asarray(star_data["source_id"], dtype=np.int64)
if "gaia_g_mag" in star_data:
    stars.catalog["Gaia G Mag"] = np.asarray(star_data["gaia_g_mag"], dtype=float)
if "detector_xpix_shifted" in star_data:
    stars.catalog["Detector Xpix Shifted"] = np.asarray(
        star_data["detector_xpix_shifted"], dtype=float
    )
if "detector_ypix_shifted" in star_data:
    stars.catalog["Detector Ypix Shifted"] = np.asarray(
        star_data["detector_ypix_shifted"], dtype=float
    )
simulator.star_data = star_data

star_df = stars.catalog
print(n_stars)
# star_df

# ---------------------------------------------------------------------------
# 导星星表规则：
# 1) 先自适应加深 Gaia 查询，直到该视场至少有 GUIDE_MIN_CATALOG_STARS 颗星或达到查询上限；
# 2) 最终只把最亮的若干颗送进仿真，受 MAX_SIM_STARS 限制；
# 3) 若即使查到 GAIA_GMAG_LIM_MAX 仍不足目标星数，则给出明确告警。
# ---------------------------------------------------------------------------
try:
    max_sim_stars = int(MAX_SIM_STARS) if MAX_SIM_STARS is not None else None
except Exception:
    max_sim_stars = None

star_selection = select_guide_star_indices(
    star_df=star_df,
    min_catalog_stars=GUIDE_MIN_CATALOG_STARS,
    max_sim_stars=max_sim_stars,
)

keep = star_selection["keep_indices"]
star_df = star_df[keep]
star_data = _slice_star_data_by_indices(star_data, keep)
stars.catalog = star_df


simulator.star_data = star_data

print(
    "[Stars] "
    f"Gaia query reached G <= {guide_query_meta['query_final_gaia_gmag_lim']:.1f}; "
    f"returned {star_selection['n_available']} stars; "
    f"selected {star_selection['n_selected']} for simulation"
    + ("" if max_sim_stars is None else f" (MAX_SIM_STARS={max_sim_stars})")
    + f"; faintest selected Kepler Mag = {star_selection['selected_faintest_mag']:.3f}."
)

if not star_selection["meets_min_count"]:
    print(
        "[Stars] WARNING: "
        f"this sky field still has only {star_selection['n_available']} real stars "
        f"inside the current 1.7 deg FOV even after relaxing Gaia to "
        f"G <= {guide_query_meta['query_final_gaia_gmag_lim']:.1f}, "
        f"so it cannot reach the requested minimum {GUIDE_MIN_CATALOG_STARS} stars."
    )

truth_index = np.arange(len(star_df), dtype=np.int64)
star_df["Truth Index"] = truth_index
star_df["Star ID"] = truth_index
star_df["FOV theta"] = np.full(len(star_df), target_star_angle, dtype=np.float64)
stars.catalog = star_df

try:
    with open(meta_path, "r", encoding="utf-8") as f:
        _meta = json.load(f)
except Exception:
    _meta = {}

_meta.update(
    {
        "n_truth_stars": int(len(star_df)),
        "field_polar_angle_rad": float(target_star_angle),
        "guide_catalog_min_star_target": int(GUIDE_MIN_CATALOG_STARS),
        "guide_catalog_available_star_count": int(star_selection["n_available"]),
        "guide_catalog_selected_star_count": int(star_selection["n_selected"]),
        "guide_catalog_selected_faintest_kepler_mag": float(
            star_selection["selected_faintest_mag"]
        ),
        "guide_catalog_meets_min_star_target": bool(star_selection["meets_min_count"]),
    }
)

with open(meta_path, "w", encoding="utf-8") as f:
    json.dump(_meta, f, ensure_ascii=False, indent=2)


def store_new_dynamic_params(self):

    self.data_generation_manager.store_variant_dynamic_parameters(
        self.time,
        self.n_raw_frames,
        self.dynamic_param_config,
        self.variant_manager,
        self.observatory,
        sim_config=self.sim_config,
        target_star_data=self.star_data,
        will_convole_jitter=self.convole_jitter,
        compute_device=self.compute_device,
        torch_dtype=self.torch_dtype,
    )


# 为保证跨不同天区中心时 motion/MD/thermal 等随机序列完全一致：
# 在生成动态参数前再次重置 RNG（星场生成的 random offset 不会影响到这里）。
if RESEED_BEFORE_DYNAMIC_PARAMS:
    set_global_seed(GLOBAL_SEED)

store_new_dynamic_params(simulator)

psf_model.build()

psf_model.is_built = True

"""
- 9 min run cost for 1k detfield_id with 6 scope and 1 variant and max_bkg_star_mag=15.0
- 9 min run cost for 1k det with 1 scope and 1 variant and max_bkg_star_mag=18.0

-----------

Part of the reason it takes so long is that the PSF of each star is evaluated across all subpixels of the entire image. This is
highly inefficient since 99.999% of the star flux is within a 5 pixel radius circle of the centroid. Photsim was not designed to
run in this way.
"""

# ---------------------------------------------------------------------------
# 输出：星表与图像（流式保存）
# ---------------------------------------------------------------------------

# 保存星表（本批天区中心），方便你后续做质心/星匹配/导航星筛选。
try:
    star_df.write(opj(BATCH_DIR, "stars.ecsv"), overwrite=True)
except Exception:
    # 少数环境可能没有 ECSV writer；兜底保存为 CSV。
    star_df.to_pandas().to_csv(opj(BATCH_DIR, "stars.csv"), index=False)

n_coadd_frames = int(actor_config["n_coadd_frames"])
batch_size = max(1, int(FRAME_BATCH_SIZE))

# coadd 帧时间轴（单位：秒）
cadence_s = float(simulator.time_manager.sim_frame_duration.to(u.s).value)

print(f"[Output] run_dir={RUN_DIR}")
print(f"[Output] batch_dir={BATCH_DIR}")
print(f"[Output] frames_dir={FRAMES_DIR}")
print(f"[Frames] n_coadd_frames={n_coadd_frames}  batch_size={batch_size}")

from photsim6.plot import plot_images

preview_saved = False

if STREAM_SAVE_FRAMES:

    for start in trange(0, n_coadd_frames, batch_size):
        stop = min(start + batch_size, n_coadd_frames)
        batch_range = range(start, stop)

        # 目前 Telescope Count 被固定为 1；这里写成循环以防你后续改回多镜筒。
        for scope_id, telescope in enumerate(simulator.observatory.telescopes):

            image_generator = ImageProcessing(
                config=actor_config,
                coadd_slice_range=batch_range,
                stars=star_df,
                telescope=telescope,
                psf_model=psf_model,
                variant_manager=simulator.variant_manager,
                cloning_info=simulator.image_gen_process_config["cloning_info"],
                groups_per_step=simulator.image_gen_process_config["groups_per_step"],
                transit_rel_flux=transit_rel_flux,
            )

            variant_data_slice = image_generator.process_all_steps()

            variant_ids = np.array(sorted(variant_data_slice.keys()), dtype=int)
            batch_images = torch.stack(
                [variant_data_slice[int(vid)] for vid in variant_ids.tolist()],
                dim=0,
            )
            truth_payload = image_generator.build_truth_payload(variant_ids)

            batch_time_s = (
                np.arange(start, stop, dtype=np.float64) * cadence_s
            ).astype(np.float64)

            out_npz = opj(
                FRAMES_DIR, f"scope{scope_id}_coadd_{start:06d}_{stop-1:06d}.npz"
            )
            np.savez_compressed(
                out_npz,
                images=batch_images.cpu().numpy().astype(np.float32),
                variant_ids=variant_ids,
                coadd_start=np.int64(start),
                coadd_stop=np.int64(stop),
                time_s=batch_time_s,
                cadence_s=np.float64(cadence_s),
                unit="electron_or_adu",  # photsim6 的单位链可能随配置而变，这里先不强行假定。
                **truth_payload,
            )

            # 保存一个预览 PNG（第一批次的第一帧、variant=0）
            if not preview_saved:
                try:
                    first_frame = batch_images[0, 0].cpu().numpy()
                    if first_frame.size == 0:
                        raise ValueError("empty preview frame")

                    fig, ax = plt.subplots(figsize=(10, 10))
                    show_frame = np.log10(np.clip(first_frame, 1e-6, None))
                    ax.imshow(show_frame, origin="lower", cmap="gray")
                    ax.set_title("Preview Frame (log10 scale)")
                    ax.set_xlabel("x [pix]")
                    ax.set_ylabel("y [pix]")
                    fig.tight_layout()
                    fig.savefig(
                        opj(
                            BATCH_DIR,
                            f"preview_batch{FIELD_CENTER_INDEX}_"
                            f"ra{FIELD_CENTER_RA_DEG:.4f}_dec{FIELD_CENTER_DEC_DEG:.4f}_"
                            f"frame{n_pixels}pix_FOV{FOV_DEG:.1f}deg_psf{field_id}.png",
                        )
                    )
                except Exception as exc:
                    print(f"[Preview] skipped preview image: {exc}")
                finally:
                    plt.close("all")
                    preview_saved = True

            # 主动释放引用，避免 Python/torch 缓存导致内存缓慢上涨
            del variant_data_slice
            del batch_images
            del image_generator

            if torch.cuda.is_available():
                torch.cuda.empty_cache()

else:
    # 兼容：如果你真的想一次性生成（不建议 10k 帧），保留原始接口。
    images = generate_images_for_star(
        star_df,
        config=actor_config,
        psf_model=psf_model,
        variant_manager=simulator.variant_manager,
        observatory=simulator.observatory,
        image_cloning_info=simulator.image_gen_process_config["cloning_info"],
        image_groups_per_step=simulator.image_gen_process_config["groups_per_step"],
        transit_rel_flux=transit_rel_flux,
    )

    images_npz = opj(BATCH_DIR, "images_fullcube.npz")
    np.savez_compressed(images_npz, images=images.cpu().numpy())
